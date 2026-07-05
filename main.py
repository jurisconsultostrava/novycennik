# -*- coding: utf-8 -*-
"""Cenotvorba moje-zlato.cz – webová aplikace (FastAPI).
Vstup: export ceníku (CSV), volitelně productsComplete XML (kategorie), volitelně ruční PDF.
Výstup: Shoptet import CSV/XML: code,pairCode,name,guid,price,purchasePrice,Category,variant:Váha
"""
import os, io, csv, json, time, logging, asyncio
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, Response
import core
import novinky as nv
import obohaceni as ob
import openpyxl

log = logging.getLogger("cenotvorba")
app = FastAPI(title="Cenotvorba moje-zlato.cz")
APP_TOKEN = os.environ.get("APP_TOKEN", "")          # doporučeno nastavit na Railway
MAPPING_PATH = os.environ.get("MAPPING_PATH", "mapping.json")
# --- automatický feed (cron uvnitř aplikace) ---
FEED_TOKEN   = os.environ.get("FEED_TOKEN", "")
FEED_METALS  = os.environ.get("FEED_METALS", "gold").split(",")
FEED_MIN     = int(os.environ.get("FEED_INTERVAL_MIN", "60"))
F_MARGIN     = float(os.environ.get("MARGIN_PCT", "1.25"))
F_BANDS      = json.loads(os.environ.get("MARGIN_BANDS", "null") or "null")
F_ROUND      = int(os.environ.get("ROUNDING", "1"))
F_QTY        = int(os.environ.get("QTY", "1"))
_feed = {"xml": None, "ts": 0, "meta": {}, "last_error": None, "last_ok_ts": 0}

_cache = {}   # {metal: (ts, items)}
TTL = 600

def auth(x_token):
    if APP_TOKEN and x_token != APP_TOKEN:
        raise HTTPException(401, "Chybný nebo chybějící token (hlavička X-Token)")

def get_mapping():
    with open(MAPPING_PATH, encoding="utf-8") as f:
        return json.load(f)

def get_catalog(metals, pdf_uploads):
    """Katalog položek: z ručně nahraných PDF (přednost), jinak fetch s cache."""
    items = []
    if pdf_uploads:
        for b in pdf_uploads:
            items += core.parse_catalog(b)
        return items, "ruční PDF"
    for m in metals:
        c = _cache.get(m)
        if not c or time.time() - c[0] > TTL:
            _cache[m] = (time.time(), core.parse_catalog(core.fetch_stonex_pdf(m)))
        items += _cache[m][1]
    return items, "on-line fetch"

def read_cenik(b: bytes):
    txt = b.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(txt)))

def params_common(spot_mode, spot_manual, fx_mode, fx_manual, catalog):
    if spot_mode == "manual":
        spot, spot_info = float(spot_manual), "ručně"
    else:
        spot, n = core.implied_spot(catalog); spot_info = f"implikovaný z PDF (n={n})"
    if fx_mode == "manual":
        fx, fx_info = float(fx_manual), "ručně"
    else:
        fx, d = core.cnb_eur_czk(); fx_info = f"ČNB {d}"
    return spot, spot_info, fx, fx_info

async def collect(cenik, xml, pdfs):
    cen = read_cenik(await cenik.read())
    cats = core.load_categories_from_xml(await xml.read()) if xml else {}
    pdf_bytes = [await p.read() for p in (pdfs or []) if p.filename]
    return cen, cats, pdf_bytes

def run(cen, cats, pdf_bytes, metals, spot_mode, spot_manual, fx_mode, fx_manual,
        margin, bands_json, rounding, qty):
    catalog, src = get_catalog([m for m in metals.split(",") if m], pdf_bytes)
    spot, spot_info, fx, fx_info = params_common(spot_mode, spot_manual, fx_mode, fx_manual, catalog)
    bands = json.loads(bands_json) if bands_json else None
    rows, skipped = core.compute_rows(cen, get_mapping(), catalog, spot, fx,
                                      float(margin), bands, int(rounding), int(qty), cats)
    meta = dict(katalog=f"{len(catalog)} položek ({src})", spot=f"{spot} €/g ({spot_info})",
                kurz=f"{fx} CZK/EUR ({fx_info})", oceneno=len(rows), preskoceno=len(skipped))
    return rows, skipped, meta

@app.post("/api/preview")
async def preview(cenik: UploadFile = File(...), xml: UploadFile = File(None),
                  pdfs: list[UploadFile] = File(None),
                  metals: str = Form("gold"), spot_mode: str = Form("pdf"),
                  spot_manual: str = Form("0"), fx_mode: str = Form("cnb"),
                  fx_manual: str = Form("0"), margin: str = Form("1.25"),
                  bands: str = Form(""), rounding: str = Form("1"),
                  qty: str = Form("1"), x_token: str = Header("")):
    auth(x_token)
    cen, cats, pdf_bytes = await collect(cenik, xml, pdfs)
    rows, skipped, meta = run(cen, cats, pdf_bytes, metals, spot_mode, spot_manual,
                              fx_mode, fx_manual, margin, bands, rounding, qty)
    return JSONResponse({"meta": meta, "rows": rows,
                         "skipped": [{"code": s.get("code"), "name": s.get("name"),
                                      "duvod": s.get("_duvod")} for s in skipped]})

@app.post("/api/export")
async def export(fmt: str = Form("csv"), cenik: UploadFile = File(...),
                 xml: UploadFile = File(None), pdfs: list[UploadFile] = File(None),
                 metals: str = Form("gold"), spot_mode: str = Form("pdf"),
                 spot_manual: str = Form("0"), fx_mode: str = Form("cnb"),
                 fx_manual: str = Form("0"), margin: str = Form("1.25"),
                 bands: str = Form(""), rounding: str = Form("1"),
                 qty: str = Form("1"), x_token: str = Header("")):
    auth(x_token)
    cen, cats, pdf_bytes = await collect(cenik, xml, pdfs)
    rows, _, _ = run(cen, cats, pdf_bytes, metals, spot_mode, spot_manual,
                     fx_mode, fx_manual, margin, bands, rounding, qty)
    stamp = time.strftime("%Y%m%d-%H%M")
    if fmt == "xml":
        return Response(core.export_xml(rows), media_type="application/xml",
            headers={"Content-Disposition": f"attachment; filename=ceny-{stamp}.xml"})
    return Response(core.export_csv(rows), media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=ceny-{stamp}.csv"})


@app.post("/api/novinky")
async def novinky_ep(fmt: str = Form("preview"), metals: str = Form("gold"),
                     pdfs: list[UploadFile] = File(None),
                     spot_mode: str = Form("pdf"), spot_manual: str = Form("0"),
                     fx_mode: str = Form("cnb"), fx_manual: str = Form("0"),
                     margin: str = Form("1.25"), rounding: str = Form("1"),
                     limit: str = Form("0"), x_token: str = Header("")):
    """Nespárované položky dodavatele s prodejní cenou (bez sold out)
    jako nové produkty vč. Product number, Mint, dostupnosti a obrázku."""
    auth(x_token)
    pdf_bytes=[await p.read() for p in (pdfs or []) if p.filename]
    catalog,_=get_catalog([m for m in metals.split(",") if m], pdf_bytes)
    spot, si, fx, fi = params_common(spot_mode, spot_manual, fx_mode, fx_manual, catalog)
    rows_all, errs = [], []
    for m in metals.split(","):
        if not m: continue
        r,e = nv.collect_new(get_mapping(), m, spot, fx, float(margin),
                             int(rounding), limit=int(limit) or None)
        rows_all+=r; errs+=e
    if fmt=="csv":
        stamp=time.strftime("%Y%m%d-%H%M")
        return Response(nv.export_new_csv(rows_all), media_type="text/csv; charset=utf-8",
          headers={"Content-Disposition": f"attachment; filename=nove-produkty-{stamp}.csv"})
    return JSONResponse({"meta":{"spot":f"{spot} €/g ({si})","kurz":f"{fx} ({fi})",
      "novinek":len(rows_all),"chyb":len(errs)},"rows":rows_all,
      "errs":[{"name":a,"url":b,"err":c} for a,b,c in errs]})


@app.post("/api/obohatit")
async def obohatit(soubor: UploadFile = File(...), categories: UploadFile = File(None),
                   metals: str = Form("gold,silver,platinum"),
                   margin: str = Form("1.25"), rounding: str = Form("1"),
                   qty: str = Form("1"), fmt: str = Form("csv"),
                   x_token: str = Header("")):
    """Obohatí nahraný XLSX/CSV (code=StoneX Product number + name) o data z webu
    StoneX: price, purchasePrice, manufacturer, availability, image, variant:Váha
    a Category (dle nahraného categories.csv, jinak vestavěného). guid zůstává
    prázdný (přiděluje Shoptet novému produktu)."""
    auth(x_token)
    raw=await soubor.read()
    rows=[]
    if soubor.filename.lower().endswith((".xlsx",".xlsm")):
        import io as _io
        wb=openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
        ws=wb.active
        hdr=[str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1,max_row=1))]
        ci=hdr.index("code") if "code" in hdr else 0
        ni=hdr.index("name") if "name" in hdr else 1
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[ci] is None: continue
            rows.append({"code":r[ci],"name":r[ni] if ni<len(r) else ""})
    else:
        import csv as _csv, io as _io
        rd=_csv.DictReader(_io.StringIO(raw.decode("utf-8-sig")))
        for r in rd: rows.append({"code":r.get("code"),"name":r.get("name","")})
    ctext = (await categories.read()).decode("utf-8-sig") if categories else \
            open("categories.csv",encoding="utf-8-sig").read()
    mets=tuple(m for m in metals.split(",") if m)
    out,errs = ob.enrich(rows, ctext, metals=mets, margin_pct=float(margin),
                         rounding=int(rounding), qty=int(qty))
    if fmt=="csv":
        stamp=time.strftime("%Y%m%d-%H%M")
        return Response(ob.export_enriched_csv(out), media_type="text/csv; charset=utf-8",
          headers={"Content-Disposition": f"attachment; filename=obohaceno-{stamp}.csv"})
    return JSONResponse({"celkem":len(out),"chyb":len(errs),
        "rows":[{k:v for k,v in r.items() if not k.startswith("_")} |
                {"catconf":r.get("_catconf"),"zdroj":r.get("_zdroj")} for r in out]})

@app.get("/api/mapping")
def mapping_get(x_token: str = Header("")):
    auth(x_token); return get_mapping()

@app.post("/api/mapping")
async def mapping_set(body: dict, x_token: str = Header("")):
    auth(x_token)
    mp = get_mapping(); mp.update(body)
    with open(MAPPING_PATH, "w", encoding="utf-8") as f:
        json.dump(mp, f, ensure_ascii=False, indent=1)
    return {"ok": True, "polozek": len(mp),
            "pozn": "Souborový zápis je na Railway efemérní – trvalé změny commitněte do repa."}


def _regen_feed():
    catalog=[]
    try:
        for met in FEED_METALS:
            if met: catalog += core.parse_catalog(core.fetch_stonex_pdf(met.strip()))
    except Exception as e:
        _feed["last_error"] = f"{time.strftime('%Y-%m-%d %H:%M:%S')} UTC: stažení/parse selhalo: {e}"
        raise
    spot, n = core.implied_spot(catalog)
    fx, d = core.cnb_eur_czk()
    mp = get_mapping()
    pseudo=[{"code": c} for c in mp]           # ceník není třeba: stačí kódy z mapy
    rows, skipped = core.compute_rows(pseudo, mp, catalog, spot, fx,
                                      F_MARGIN, F_BANDS, F_ROUND, F_QTY, {})
    meta=(f"generováno {time.strftime('%Y-%m-%d %H:%M:%S')} UTC · spot {spot} EUR/g "
          f"(implikovaný, n={n}) · kurz CNB {fx} ({d}) · marže {F_MARGIN}% · "
          f"položek {len(rows)} · přeskočeno {len(skipped)}")
    _feed["xml"]=core.export_feed_xml(rows, meta)
    _feed["ts"]=time.time()
    _feed["meta"]={"spot":spot,"fx":fx,"rows":len(rows),
                   "skipped":[s.get("code") for s in skipped],"info":meta}
    _feed["last_ok_ts"]=time.time(); _feed["last_error"]=None
    log.info("feed přegenerován: %s", meta)

@app.on_event("startup")
async def _scheduler():
    async def loop():
        while True:
            try:
                await asyncio.to_thread(_regen_feed)
            except Exception as e:
                log.exception("regenerace feedu selhala: %s", e)
            await asyncio.sleep(FEED_MIN*60)
    asyncio.create_task(loop())

@app.get("/feed.xml")
def feed_xml(token: str = ""):
    if FEED_TOKEN and token != FEED_TOKEN:
        raise HTTPException(401, "chybný token (?token=...)")
    if _feed["xml"] is None or time.time()-_feed["ts"] > FEED_MIN*60*2:
        _regen_feed()                     # samoléčba: na vyžádání, je-li cache stará
    return Response(_feed["xml"], media_type="application/xml")

@app.get("/feed/status")
def feed_status(x_token: str = Header("")):
    auth(x_token)
    now=time.time()
    age = round((now-_feed["ts"])/60,1) if _feed["ts"] else None
    ok_age = round((now-_feed["last_ok_ts"])/60,1) if _feed["last_ok_ts"] else None
    stale = (age is None) or (age > FEED_MIN*2)
    warn=None
    if _feed["last_error"]:
        warn=f"POZOR: poslední stažení z StoneX selhalo – {_feed['last_error']}. Feed slouží poslední úspěšná data."
    elif stale:
        warn="POZOR: feed je zastaralý (starší než 2× interval)."
    return {"stav": "CHYBA" if _feed["last_error"] else ("ZASTARALÝ" if stale else "OK"),
            "stari_feedu_min": age, "od_posledniho_uspechu_min": ok_age,
            "interval_min": FEED_MIN, "posledni_chyba": _feed["last_error"],
            "varovani": warn, **_feed["meta"]}

@app.get("/", response_class=HTMLResponse)
def ui():
    return UI

UI = """<!doctype html><html lang=cs><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1"><title>Cenotvorba moje-zlato.cz</title>
<style>
body{font-family:system-ui,Segoe UI,Arial;margin:0;background:#141f17;color:#e7e1cd}
header{padding:18px 24px;border-bottom:1px solid #c9a24b55}
h1{font-size:18px;margin:0;letter-spacing:.05em}h1 b{color:#c9a24b}
main{max-width:1100px;margin:0 auto;padding:20px}
fieldset{border:1px solid #c9a24b44;border-radius:10px;margin:0 0 16px;padding:14px 16px}
legend{color:#c9a24b;font-weight:700;font-size:13px;letter-spacing:.08em}
label{display:inline-block;margin:4px 14px 4px 0;font-size:13px}
input,select{background:#1c2a1f;color:#e7e1cd;border:1px solid #c9a24b55;border-radius:6px;padding:6px 8px}
input[type=file]{border:0;padding:2px}
button{background:linear-gradient(180deg,#e0bd6a,#c9a24b 55%,#a8842f);color:#231a06;border:0;
 border-radius:8px;padding:10px 18px;font-weight:800;cursor:pointer;margin-right:10px}
#meta{font-size:13px;color:#a8b3a2;margin:10px 0}
table{border-collapse:collapse;width:100%;font-size:12px;margin-top:10px}
th,td{border-bottom:1px solid #ffffff22;padding:5px 7px;text-align:left}
th{color:#c9a24b}.up{color:#8fd18f}.down{color:#e09090}.warn{color:#e0bd6a}
#err{color:#e09090;white-space:pre-wrap}
</style></head><body>
<header><h1>CENOTVORBA <b>moje-zlato.cz</b> · StoneX → Shoptet</h1></header><main>
<fieldset><legend>PŘÍSTUP</legend>
<label>Token: <input id=tok type=password placeholder="X-Token (je-li nastaven)"></label></fieldset>
<fieldset><legend>VSTUPY</legend>
<label>Ceník CSV (export Shoptet): <input id=cenik type=file accept=.csv required></label><br>
<label>productsComplete XML (kategorie, volitelné): <input id=xml type=file accept=.xml></label><br>
<label>StoneX PDF ručně (volitelné, má přednost před on-line stažením): <input id=pdfs type=file accept=.pdf multiple></label>
</fieldset>
<fieldset><legend>PARAMETRY</legend>
<label>Kovy: <label><input type=checkbox class=met value=gold checked>zlato</label>
<label><input type=checkbox class=met value=silver>stříbro</label>
<label><input type=checkbox class=met value=platinum>platina</label>
<label><input type=checkbox class=met value=palladium>palladium</label></label><br>
<label>Spot: <select id=spotm><option value=pdf>implikovaný z PDF (doporučeno)</option>
<option value=manual>ručně</option></select>
<input id=spotv type=number step=0.01 placeholder="€/g" style="width:90px"></label>
<label>Kurz: <select id=fxm><option value=cnb>ČNB automaticky</option>
<option value=manual>ručně</option></select>
<input id=fxv type=number step=0.001 placeholder="CZK/EUR" style="width:90px"></label><br>
<label>Marže %: <input id=margin type=number step=0.05 value=1.25 style="width:80px"></label>
<label>Pásma (JSON, volit.): <input id=bands placeholder='[{"max_g":10,"pct":5}]' style="width:240px"></label>
<label>Zaokrouhlení: <select id=round><option>1</option><option>10</option><option>100</option></select> Kč</label>
<label>Odběr ks (pásmo prémie): <input id=qty type=number value=1 style="width:60px"></label>
</fieldset>
<fieldset><legend>NOVINKY OD DODAVATELE</legend>
<label>Limit položek (0 = vše): <input id=nlimit type=number value=0 style="width:70px"></label>
<button onclick=novinky('preview')>Náhled novinek</button>
<button onclick=novinky('csv')>Stáhnout CSV novinek</button>
<p style="font-size:12px;color:#a8b3a2;margin:6px 0 0">Položky dodavatele, které e-shop nemá
(dle mapy párování), s dostupnou prodejní cenou; sold out se vynechává. Přenáší se:
Product number → code, Mint → výrobce, hmotnost, dostupnost a CDN obrázek.
Sloupec Category zůstává prázdný k doplnění před importem.</p></fieldset>
<fieldset><legend>OBOHACENÍ NESPÁROVANÝCH (XLSX/CSV → nové produkty)</legend>
<label>Soubor (code + name): <input id=obsoubor type=file accept=.xlsx,.csv></label>
<label>categories.csv (volit., jinak vestavěný): <input id=obcats type=file accept=.csv></label><br>
<button onclick=obohatit('preview')>Náhled obohacení</button>
<button onclick=obohatit('csv')>Stáhnout obohacené CSV</button>
<p style="font-size:12px;color:#a8b3a2;margin:6px 0 0">Doplní price, purchasePrice, manufacturer(Mint),
availability, image(CDN) a variant:Váha z veřejného webu StoneX (párování přes Product number = code);
Category se určí z categories.csv dle názvu. guid zůstává prázdný – přidělí Shoptet. Stahování detailů
trvá; sold out se označí. Běží spolehlivě jen z nasazení (Railway), ne z lokálního prostředí za Cloudflare.</p></fieldset>
<button onclick=go('preview')>Náhled</button>
<button onclick=go('csv')>Stáhnout CSV</button>
<button onclick=go('xml')>Stáhnout XML</button>
<div id=meta></div><div id=err></div><div id=out></div>
<script>
function fd(){const f=new FormData();
 f.append('cenik',cenik.files[0]);
 if(xml.files[0])f.append('xml',xml.files[0]);
 for(const p of pdfs.files)f.append('pdfs',p);
 f.append('metals',[...document.querySelectorAll('.met:checked')].map(x=>x.value).join(','));
 f.append('spot_mode',spotm.value);f.append('spot_manual',spotv.value||'0');
 f.append('fx_mode',fxm.value);f.append('fx_manual',fxv.value||'0');
 f.append('margin',margin.value);f.append('bands',bands.value);
 f.append('rounding',document.getElementById('round').value);f.append('qty',qty.value);
 return f}
async function go(mode){err.textContent='';out.innerHTML='';meta.textContent='Pracuji…';
 if(!cenik.files[0]){err.textContent='Nahrajte ceník CSV.';meta.textContent='';return}
 const f=fd();const h={'X-Token':tok.value};
 try{
  if(mode=='preview'){
   const r=await fetch('/api/preview',{method:'POST',body:f,headers:h});
   if(!r.ok)throw new Error(await r.text());
   const d=await r.json();
   meta.textContent=`Katalog: ${d.meta.katalog} · Spot: ${d.meta.spot} · Kurz: ${d.meta.kurz} · Oceněno: ${d.meta.oceneno} · Přeskočeno: ${d.meta.preskoceno}`;
   let h1='<table><tr><th>code</th><th>název</th><th>StoneX</th><th>váha g</th><th>prémie €</th><th>nákup CZK</th><th>stará</th><th>NOVÁ cena</th><th>Δ%</th><th>shoda</th></tr>';
   for(const r of d.rows){const o=parseFloat((r._stara_cena||'0').replace(',','.'));
    const dl=o?((r.price-o)/o*100):0;const cls=dl>0?'up':dl<0?'down':'';
    h1+=`<tr><td>${r.code}</td><td>${r.name}</td><td>${r._stonex}</td><td>${r._wg}</td><td>${r._prem_eur}</td><td>${r.purchasePrice}</td><td>${r._stara_cena}</td><td><b>${r.price}</b></td><td class=${cls}>${o?dl.toFixed(2):''}</td><td class=${r._shoda=='přesná'?'':'warn'}>${r._shoda}</td></tr>`}
   h1+='</table>';
   if(d.skipped.length){h1+=`<p class=warn>Přeskočeno ${d.skipped.length}:</p><table><tr><th>code</th><th>název</th><th>důvod</th></tr>`;
    for(const s of d.skipped)h1+=`<tr><td>${s.code||''}</td><td>${s.name||''}</td><td>${s.duvod}</td></tr>`;h1+='</table>'}
   out.innerHTML=h1;
  }else{
   f.append('fmt',mode);
   const r=await fetch('/api/export',{method:'POST',body:f,headers:h});
   if(!r.ok)throw new Error(await r.text());
   const b=await r.blob();const a=document.createElement('a');
   a.href=URL.createObjectURL(b);
   a.download=(r.headers.get('Content-Disposition')||'').split('filename=')[1]||('ceny.'+mode);
   a.click();meta.textContent='Soubor stažen.';
  }
 }catch(e){err.textContent=e.message;meta.textContent=''}
}

async function novinky(mode){err.textContent='';out.innerHTML='';meta.textContent='Stahuji detaily produktů, může to trvat i minuty…';
 const f=new FormData();
 for(const p of pdfs.files)f.append('pdfs',p);
 f.append('metals',[...document.querySelectorAll('.met:checked')].map(x=>x.value).join(','));
 f.append('spot_mode',spotm.value);f.append('spot_manual',spotv.value||'0');
 f.append('fx_mode',fxm.value);f.append('fx_manual',fxv.value||'0');
 f.append('margin',margin.value);f.append('rounding',document.getElementById('round').value);
 f.append('limit',nlimit.value||'0');f.append('fmt',mode=='csv'?'csv':'preview');
 try{
  const r=await fetch('/api/novinky',{method:'POST',body:f,headers:{'X-Token':tok.value}});
  if(!r.ok)throw new Error(await r.text());
  if(mode=='csv'){const b=await r.blob();const a=document.createElement('a');
   a.href=URL.createObjectURL(b);a.download='nove-produkty.csv';a.click();
   meta.textContent='Soubor stažen.';return}
  const d=await r.json();
  meta.textContent=`Spot: ${d.meta.spot} · Kurz: ${d.meta.kurz} · Novinek: ${d.meta.novinek} · Chyb: ${d.meta.chyb}`;
  let h='<table><tr><th>code</th><th>název</th><th>výrobce (Mint)</th><th>váha g</th><th>dostupnost</th><th>nákup CZK</th><th>cena CZK</th><th>obrázek</th></tr>';
  for(const x of d.rows)h+=`<tr><td>${x.code}</td><td><a href="${x.url}" target=_blank style="color:#c9a24b">${x.name}</a></td><td>${x.manufacturer}</td><td>${x['variant:Váha']}</td><td>${x.availability}</td><td>${x.purchasePrice}</td><td><b>${x.price}</b></td><td>${x.image?'✓':'—'}</td></tr>`;
  h+='</table>';
  if(d.errs.length){h+='<p class=warn>Chyby:</p>';for(const e of d.errs)h+=`<div class=warn style="font-size:12px">${e.name}: ${e.err}</div>`}
  out.innerHTML=h;
 }catch(e){err.textContent=e.message;meta.textContent=''}
}

async function obohatit(mode){err.textContent='';out.innerHTML='';meta.textContent='Stahuji detaily z StoneX, může to trvat minuty…';
 if(!obsoubor.files[0]){err.textContent='Nahrajte XLSX/CSV se sloupci code a name.';meta.textContent='';return}
 const f=new FormData();f.append('soubor',obsoubor.files[0]);
 if(obcats.files[0])f.append('categories',obcats.files[0]);
 f.append('metals',[...document.querySelectorAll('.met:checked')].map(x=>x.value).join(',')||'gold,silver,platinum');
 f.append('margin',margin.value);f.append('rounding',document.getElementById('round').value);
 f.append('qty',qty.value);f.append('fmt',mode=='csv'?'csv':'preview');
 try{
  const r=await fetch('/api/obohatit',{method:'POST',body:f,headers:{'X-Token':tok.value}});
  if(!r.ok)throw new Error(await r.text());
  if(mode=='csv'){const b=await r.blob();const a=document.createElement('a');
   a.href=URL.createObjectURL(b);a.download='obohaceno.csv';a.click();meta.textContent='Staženo.';return}
  const d=await r.json();
  meta.textContent=`Celkem ${d.celkem} · s chybou ${d.chyb}`;
  let h='<table><tr><th>code</th><th>název</th><th>výrobce</th><th>váha</th><th>Category</th><th>dostupnost</th><th>nákup</th><th>cena</th><th>obr.</th><th>zdroj/chyba</th></tr>';
  for(const x of d.rows){const okimg=x.image?'✓':'—';const bad=x.price===''?'warn':'';
   h+=`<tr class=${bad}><td>${x.code}</td><td>${x.name}</td><td>${x.manufacturer}</td><td>${x['variant:Váha']}</td><td>${x.Category}<br><small>${x.catconf||''}</small></td><td>${x.availability}</td><td>${x.purchasePrice}</td><td><b>${x.price}</b></td><td>${okimg}</td><td><small>${(x.zdroj||'').slice(0,60)}</small></td></tr>`}
  h+='</table>';out.innerHTML=h;
 }catch(e){err.textContent=e.message;meta.textContent=''}
}
</script></main></body></html>"""
